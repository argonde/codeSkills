#! /usr/bin/python3
# Developer: Ruben López
# -*- coding: utf-8 -*-
import os
import subprocess
import pandas as pd
import openpyxl as xl
from pathlib import Path
import pyodbc


# Available functions
# Correct column's name by position
def correct(col_, head_):
    for flag_, value_ in head_.items():
        # match column position
        if col_ == value_:
            return flag_


# Search in a dictionary for the value's key
def get_key(val_, dict_):
    for key_, value_ in dict_.items():
        # match the values and then get the key
        if str(val_) == str(value_):
            return key_


# Find the index of a value in a list
def find_indeks(val_, list_):
    indeks_ = list_.index(val_)
    return indeks_


# Calculate the size of the excel spreadsheet
def xls_size(wb_):
    n_cols_ = 0
    max_n_cols_ = 0
    # Create dictionary and count column duplicates
    # {cellValue: #}
    count_ = {}
    # Get the active tab...
    sheet_ = wb_.active
    max_cols_ = sheet_.max_column
    # Loop and use all cell values
    rows_ = [[parse_xls_cells_in_row(cell_) for cell_ in row_] for row_ in sheet_.iter_rows()]
    n_rows_ = len(rows_)
    for col_ in range(1, max_cols_):
        sc_ = sheet_.cell(row=1, column=col_).value
        count_.setdefault(sc_, 1)
        # Count empty cells
        if sc_ is not None:
            max_n_cols_ += 1
        # Count cell value, if it repeats
        elif sc_ in count_:
            count_[sc_] = count_[sc_] + 1
        # Count all cells
        else:
            pass
        n_cols_ += 1
    return max_n_cols_, n_cols_, n_rows_, count_


# Select a matching empty global variable to assign its value
def select(d_):
    k_list_ = []
    # match existing empty variable with the hard-coded dictionary
    list_of_globals_ = globals()
    for k_ in list_of_globals_:
        if k_.startswith(d_):
            k_list_.append([k_, list_of_globals_[k_]])
    return k_list_


# Parse excel cells from a given row
def parse_xls_cells_in_row(cell_):
    cell_value_ = cell_.value
    # use only cells that are not empty
    if cell_value_ is not None:
        return cell_value_ or ''


# Parse an Excel file
def parse_file(file_):
    # Get the name
    if file_[-5:] == ".xlsx":
        f_ = os.path.basename(file_[:-5])
        try:
            # Open the Excel file
            wb_ = xl.load_workbook(file_)
            # Get all tabs in the excel sheet...
            # sheet_ = wb_.active
            if len(wb_.sheetnames) > 1:
                for w_name in wb_.sheetnames:
                    sheet_ = wb_[w_name]
                    # Parse the data in the tab
                    rows_ = [[parse_xls_cells_in_row(cell_) for cell_ in row_] for row_ in sheet_.iter_rows()]
                    # Proceed only if the file is not empty
                    if rows_ is not None:
                        # Trim subsequent empty rows
                        last_row_index_ = max(idx_ for idx_, row_ in enumerate(rows_) if any(val_ for val_ in row_))
                        rows_ = rows_[:last_row_index_ + 1]
                        # Debug info
                        # print("Row:", rows_)
                        # All rows must have the same number of columns
                        last_col_index_ = max(max(idx_ for idx_, val_ in enumerate(row_) if val_) for row_ in rows_)
                        padding_ = [''] * last_col_index_
                        # Assign the resulting list to a variable
                        rows_ = [(row_ + padding_)[:last_col_index_ + 1] for row_ in rows_]
                        return rows_, f_, wb_.sheetnames
                    else:
                        rows_ = 0
                        return rows_, f_, wb_.sheetnames
            else:
                sheet_ = wb_.active
                _sheet_ = str(sheet_)[12:-2]
                # Parse the data in the tab
                rows_ = [[parse_xls_cells_in_row(cell_) for cell_ in row_] for row_ in sheet_.iter_rows()]
                # Trim subsequent empty rows
                last_row_index_ = max(i_ for i_, row_ in enumerate(rows_) if any(val_ for val_ in row_))
                rows_ = rows_[:last_row_index_ + 1]
                # Debug info
                # print("Row:", rows_)
                # All rows must have the same number of columns
                last_col_index_ = max(max(i_ for i_, val_ in enumerate(row_) if val_) for row_ in rows_)
                padding_ = [''] * last_col_index_
                # Assign the resulting list to a variable
                rows_ = [(row_ + padding_)[:last_col_index_ + 1] for row_ in rows_]
                return rows_, f_, _sheet_
        # Take care of possible exceptions
        except FileNotFoundError:
            print("No file found")
        except IsADirectoryError:
            print("The input is not a file, but a directory")
    elif file_[:-4] == ".xls":
        f_ = os.path.basename(file_[:-4])
        try:
            # Open the Excel file
            wb_ = xl.load_workbook(file_)
            # Fetch the active tab ...
            sheet_ = wb_.active
            # Parse the data in the tab
            rows_ = [[parse_xls_cells_in_row(cell_) for cell_ in row_] for row_ in sheet_.iter_rows()]
            # Trim subsequent empty rows
            last_row_index_ = max(i_ for i_, row_ in enumerate(rows_) if any(val_ for val_ in row_))
            rows_ = rows_[:last_row_index_ + 1]
            # All rows must have the same number of columns
            last_col_index_ = max(max(i_ for i_, val_ in enumerate(row_) if val_) for row_ in rows_)
            padding_ = [''] * last_col_index_
            # Assign the resulting list to a variable
            rows_ = [(row_ + padding_)[:last_col_index_ + 1] for row_ in rows_]
            return rows_, f_, sheet_
        # Take care of possible exceptions
        except FileNotFoundError:
            print("No file found")
        except IsADirectoryError:
            print("The input is not a file, but a directory")

    # Take care of other possible file types
    else:
        raise ValueError("Unexpected file type: {}".format(file_))


# Choose the latest version of each file
def current_version(wd_):
    # {file: #}
    count_ = {}
    # {file: filePath}
    file_list_ = {}
    # {file#: full filePath}
    file_list_paths_ = {}
    # Loop through the files found in subdirectories ...
    for dir_path_, dirs_, files_ in os.walk(wd_):
        for file_ in files_:
            # ... select only files with excel extensions ...
            if file_.endswith('.xlsx') or file_.endswith('.xls'):
                fpath_ = os.path.join(dir_path_, file_)
                # ... write down the files and count their duplicates
                count_.setdefault(file_, 0)
                count_[file_] = count_[file_] + 1
                dict_paths_ = {str(file_ + str(count_[file_])): str(fpath_)}
                file_list_paths_.update(dict_paths_)
                # Loop through.xlsx files and their amount of duplicates
                for k_, v_ in count_.items():
                    # Write down the latest version in the dictionary, if duplicates of such file were found.
                    if int(v_) > 1:
                        # {time: filePath}
                        list_paths = {}
                        # Loop through the amount of duplicates ...
                        for f_ in range(1, v_):
                            path_ = Path(file_list_paths_.get(str(k_) + str(v_)))
                            time_ = path_.stat().st_mtime
                            paths_ = {time_: str(path_)}
                            list_paths.update(paths_)
                        # ... select the latest key, ...
                        key_ = max(list_paths)
                        f_ = os.path.basename(list_paths.get(key_))
                        file_dup_ = {f_: list_paths.get(key_)}
                        file_list_.update(file_dup_)
                    # ... otherwise add direct file names without duplicates
                    else:
                        path_ = file_list_paths_.get(str(k_) + str(v_))
                        f_ = os.path.basename(path_)
                        file_simp_ = {f_: str(path_)}
                        file_list_.update(file_simp_)
            #  ... select only files with .csv extension ...
            elif file_.endswith('.csv'):
                fpath_ = os.path.join(dir_path_, file_)
                # ... write down the files and count their duplicates.
                count_.setdefault(file_, 0)
                count_[file_] = count_[file_] + 1
                dict_paths_ = {str(file_ + str(count_[file_])): str(fpath_)}
                file_list_paths_.update(dict_paths_)

                # Loop through .csv files and number of duplicates.
                for k_, v_ in count_.items():
                    # Write down the latest version in the dictionary, if duplicates of such file were found.
                    if int(v_) > 1:
                        # {time: filePath}
                        list_paths = {}
                        # Loop through the amount of duplicates ...
                        for f_ in range(1, v_):
                            path_ = Path(file_list_paths_.get(str(k_) + str(v_)))
                            time_ = path_.stat().st_mtime
                            paths_ = {time_: str(path_)}
                            list_paths.update(paths_)
                        # ... select the latest key, ...
                        key_ = max(list_paths)
                        f_ = os.path.basename(list_paths.get(key_))
                        file_dup_ = {f_: list_paths.get(key_)}
                        file_list_.update(file_dup_)
                    # ... otherwise add direct file names without duplicates
                    else:
                        path_ = file_list_paths_.get(str(k_) + str(v_))
                        f_ = os.path.basename(path_)
                        file_simp_ = {f_: str(path_)}
                        file_list_.update(file_simp_)
            # ... select only files with mapinfo extensions ...
            elif file_.endswith('.TAB') or file_.endswith('.tab'):
                fpath_ = os.path.join(dir_path_, file_)
                # ... write down the files and count their duplicates.
                count_.setdefault(file_, 0)
                count_[file_] = count_[file_] + 1
                dict_paths_ = {str(file_ + str(count_[file_])): str(fpath_)}
                file_list_paths_.update(dict_paths_)

                #  Loop through .tab files and number of duplicates.
                for k_, v_ in count_.items():
                    # Write down the latest version in the dictionary, if duplicates of such file were found.
                    if int(v_) > 1:
                        # {time: filePath}
                        list_paths = {}
                        # Loop through the amount of duplicates ...
                        for f_ in range(1, v_):
                            path_ = Path(file_list_paths_.get(str(k_) + str(v_)))
                            time_ = path_.stat().st_mtime
                            paths_ = {time_: str(path_)}
                            list_paths.update(paths_)
                        # ... select the latest key, ...
                        key_ = max(list_paths)
                        f_ = os.path.basename(list_paths.get(key_))
                        file_dup_ = {f_: list_paths.get(key_)}
                        file_list_.update(file_dup_)
                    # ... otherwise add direct file names without duplicates
                    else:
                        path_ = file_list_paths_.get(str(k_) + str(v_))
                        f_ = os.path.basename(path_)
                        file_simp_ = {f_: str(path_)}
                        file_list_.update(file_simp_)
            # ... select only files with ESRI shapefile extensions ...
            elif file_.endswith('.shp') or file_.endswith('.SHP'):
                fpath_ = os.path.join(dir_path_, file_)
                # ... write down the files and count their duplicates.
                count_.setdefault(file_, 0)
                count_[file_] = count_[file_] + 1
                dict_paths_ = {str(file_ + str(count_[file_])): str(fpath_)}
                file_list_paths_.update(dict_paths_)

                #  Loop through .shp files and number of duplicates.
                for k_, v_ in count_.items():
                    # Write down the latest version in the dictionary, if duplicates of such file were found.
                    if int(v_) > 1:
                        # {time: filePath}
                        list_paths = {}
                        # Loop through the amount of duplicates ...
                        for f_ in range(1, v_):
                            path_ = Path(file_list_paths_.get(str(k_) + str(v_)))
                            time_ = path_.stat().st_mtime
                            paths_ = {time_: str(path_)}
                            list_paths.update(paths_)
                        # ... select the latest key, ...
                        key_ = max(list_paths)
                        f_ = os.path.basename(list_paths.get(key_))
                        file_dup_ = {f_: list_paths.get(key_)}
                        file_list_.update(file_dup_)
                    # ... otherwise add direct file names without duplicates
                    else:
                        path_ = file_list_paths_.get(str(k_) + str(v_))
                        f_ = os.path.basename(path_)
                        file_simp_ = {f_: str(path_)}
                        file_list_.update(file_simp_)
            # ... select only files with access database extensions ...
            elif file_.endswith('.accdb'):
                fpath_ = os.path.join(dir_path_, file_)
                # ... write down the files and count their duplicates.
                count_.setdefault(file_, 0)
                count_[file_] = count_[file_] + 1
                dict_paths_ = {str(file_ + str(count_[file_])): str(fpath_)}
                file_list_paths_.update(dict_paths_)

                #  Loop through .accdb files and number of duplicates.
                for k_, v_ in count_.items():
                    # Write down the latest version in the dictionary, if duplicates of such file were found.
                    if int(v_) > 1:
                        # {time: filePath}
                        list_paths = {}
                        # Loop through the amount of duplicates ...
                        for f_ in range(1, v_):
                            path_ = Path(file_list_paths_.get(str(k_) + str(v_)))
                            time_ = path_.stat().st_mtime
                            paths_ = {time_: str(path_)}
                            list_paths.update(paths_)
                        # ... select the latest key, ...
                        key_ = max(list_paths)
                        f_ = os.path.basename(list_paths.get(key_))
                        file_dup_ = {f_: list_paths.get(key_)}
                        file_list_.update(file_dup_)
                    # ... otherwise add direct file names without duplicates
                    else:
                        path_ = file_list_paths_.get(str(k_) + str(v_))
                        f_ = os.path.basename(path_)
                        file_simp_ = {f_: str(path_)}
                        file_list_.update(file_simp_)

    return file_list_


###################################################################################################
###################################################################################################
# Define global variables and create the needed dictionaries
out_path = r"/path/to/work/_data/_pyOutput/tmp/"
out_csv = r"/path/to/work/_data/_pyOutput/csv_out/"
in_path = r"/path/to/work/_data/_orgnl/0Samlet_data/"
file_search = ("Andexxxxxxxa.xlsx", "Angxxxxxxxxn.xlsx", "bxxxxxxxxxxxxi.shp",
               "bxxxxxxxxxxxn.shp", "bxxxxxxxxxxxi.shp", "bxxxxxxxxxxxxxxxxxxn.shp",
               "bxxxxxxxxxxxxxxxxx.shp", "bxxxxxxxxxxxxxxxn.shp", "baxxxxxxxxxxxxxxxxx.shp",
               "bxxxxxxxxxxxxxxxxxn.shp", "Baxxxxxxxxxxxxb.xlsx", "Bxxxxxxxxxxxxxx.xlsx", "Dxxxxxxxxxxxn.xlsx",
               "Dxxxxxxxxxxxxxx.xlsx", "Fxxxxxxxxxxxxxxxx.xlsx", "Fxxxxxxxx.xlsx", "Fxxxxxxxxxxxn.xlsx",
               "Fxxxxxxxxxxxxxx.xlsx", "gxxxxxxxxxxx.accdb", "gxxxxxxxxxxxx.accdb", "Ixxxxxxxxxxxxxx.xlsx",
               "Ixxxxxxxxxxxxxxxxx.xlsx", "Ixxxxxxxxxx.xlsx", "Kxxxxxxxxxxxxxx.xlsx", "Lxxxxxxxxxxxxxxx.xlsx",
               "Mxxxxxxxxxxxxxxxn.xlsx", "Maxxxxxxxxxxxxx.xlsx", "Mxxxxxxxxxxxxxxxxxxxxxxxxxxd.xlsx",
               "Mxxxxxxxxxxxxxxxxxxxxxxx.xlsx", "Mxxxxxxxxxxxe.xlsx", "Mxxxxxxxxxxxxxxxxxxxx.xlsx",
               "Mxxxxxxxxxxxxxxxxxxxxxx.xlsx", "Mxxxxxxxxxxxxxxxxxxxx.xlsx", "Nxxxxxxxxxxxxxxxxxx.xlsx",
               "Rxxxxxxxxxxxxxx.xlsx", "Rxxxxxxxxxxxxxxxxxxx.xlsx", "Sxxxxxxxxxxx.xlsx", "Uxxxxxxxxxxxxxxxxxxxxxx.xlsx",
               "Uxxxxxxxxxxxx.xlsx", "Uxxxxxxxxxxxxx.xlsx", "Uxxxxxxxxxxxxxx.xlsx", "xxxxxxxxxxxxxxx.xlsx",
               "vxxxxxxxxxxxxxxxxxxxxxxxxxxxx.shp", "vxxxxxxxxxxxxxxxxxxxxxxxxxx.xlsx",
               "vxxxxxxxxxxxxxxxxxxxxxx.xlsx", "vxxxxxxxxxxxxxxxxxxxxx.shp", "Vxxxxxxxxxxxxxxxxx.csv")
# Create dynamic variable lists for the Access database tables
gwPollutant = []
gwPollutant_baggrundsdata = []
# Set the working directory
os.chdir(in_path)
cwd_ = os.getcwd()
print("The cwd is now:", cwd_, "\n")
# Create dynamic dictionaries for setting up data frames.
empty_column_list = {}
duplicate_column_list = {}
# {filePath: row}
dfs_dict = {}
# Create blank data.frame for data
data_df = []
# Create blank data.frame for metadata
info_df = []
# Create two blank lists for column ids, and for their data, in order to later create a data.frame with custom data type
index_df_name = []
index_df_data = []
# Call the functions
# Fetch data from the path and check for file versions
file_list = current_version(cwd_)
print(file_list)
# Convert .tab to .shp, and then to .csv and load as data.frame
n = 0
for fil in file_list:
    if fil[-4:] == ".TAB" or fil[-4:] == ".tab":
        # Warn about problems with ogr2ogr translations, when it comes to Mapinfo files
        # There is no time to test, if there is trouble with other files, these seem to work fine.
        print("----------------------------------------WARNING: ----------------------------------------")
        print("At this point, Mapinfo files are poorly supported by the ogr translation drivers used here.")
        print("The output file may have lost many data rows. Copy/paste manually from the Mapinfo table to.csv/.xlsx")
        # Give a name to the output file.
        print("Converting %s to shp" % fil)
        shp_file = fil[:-4] + '.shp'
        # Define the input file.
        in_tab = file_list.get(str(fil))
        # Define the output file.
        out_shp = os.path.join(out_path, shp_file)
        print("----", out_shp, "-------------------------------------------------", "\n")
        # Convert Format (first step)
        subprocess.check_output(['ogr2ogr', '-f', "ESRI Shapefile", str(out_shp), str(in_tab)])
        csv_file = fil[:-4] + '.csv'
        # Define the intermediate output file
        out_csv_ = Path(out_csv, csv_file)
        # Convert Format (final step)
        subprocess.check_output(['ogr2ogr', '-f', "CSV", str(out_csv_), str(out_shp)])
        tab_file = pd.read_csv(out_csv_, dtype=object)
        # Check for any last irregular columns
        if 'NaN' in tab_file.head():
            to_drop = ['NaN']
            tab_file.drop(to_drop, inplace=True, axis=1)
        # Collect data.frames in series
        data_df.append(tab_file)
        # Make an overview of the location of data.frames, i.e. their different files
        df_dict_ = {os.path.basename(fil): n}
        dfs_dict.update(df_dict_)
        n += 1

    elif fil[-4:] == ".csv":
        # Define the input file.
        in_csv = file_list.get(str(fil))
        # Read pandas data csv data.frame
        csv_file = pd.read_csv(in_csv, sep=';', dtype=object)
        # Check for any last irregular columns
        if 'NaN' in csv_file.head():
            to_drop = ['NaN']
            csv_file.drop(to_drop, inplace=True, axis=1)
        # Collect data.frames in series
        data_df.append(csv_file)
        # Make an overview of the location of data.frames, i.e. their different files
        df_dict_ = {os.path.basename(fil): n}
        dfs_dict.update(df_dict_)
        n += 1

    # Convert all shp to csv and load as data.frame
    elif fil[-4:] == ".shp":
        # Give a name to the output file.
        print("Converting %s to csv" % fil)
        csv_file = str(fil[:-4]) + ".csv"
        # Define the input file.
        in_shp = file_list.get(str(fil))
        # Define the intermediate output file
        out_csv_ = os.path.join(out_csv, csv_file)
        print("----", out_csv_, "-------------------------------------------------", "\n")
        # Format converter
        subprocess.check_output(['ogr2ogr', '-f', "CSV", str(out_csv_), str(in_shp)])
        shp_file = pd.read_csv(out_csv_, dtype=object)
        # Check for any last irregular columns
        if 'NaN' in shp_file.head():
            to_drop = ['NaN']
            shp_file.drop(to_drop, inplace=True, axis=1)
        # Collect data.frames in series
        data_df.append(shp_file)
        # Make an overview of the location of data.frames, i.e. their different files
        df_dict_ = {os.path.basename(fil): n}
        dfs_dict.update(df_dict_)
        n += 1

    # Parse Excel sheets Parse excel sheets and load it as data.frame
    elif (fil[-5:] == ".xlsx" or fil[-4:] == ".xls") and fil in file_search:
        wb = xl.load_workbook(file_list.get(fil))
        if len(wb.sheetnames) > 1:
            for table_name in wb.sheetnames:
                sheet = wb[table_name]
                # Collect the values contained in the result of the parse file function
                excel, fil_, _s_ = parse_file(file_list.get(fil))
                # Create a data.frame from each data point, which corresponds to a tab file
                xls_df = pd.DataFrame(excel[1:], columns=excel[0], dtype=object)
                # Check for any last invalid columns
                for val in xls_df.columns:
                    if val is None:
                        to_drop = [None]
                        print("Invalid column in", fil, ":", to_drop, '\n')
                        xls_df.drop(to_drop, inplace=True, axis=1)
                    else:
                        continue
                # Collect data.frames in series
                data_df.append(xls_df)
                # Make an overview of the location of data.frames, i.e. their different files
                df_dict_ = {os.path.basename(fil) + '_' + str(table_name): n}
                dfs_dict.update(df_dict_)
                n += 1

        else:
            a_sheet = wb.active
            # collect the values contained in the result of the parse file function
            excel, fil_, sheet = parse_file(file_list.get(fil))
            # Create a data.frame from each data point, which corresponds to the file
            xls_df = pd.DataFrame(excel[1:], columns=excel[0], dtype=object)
            # Check for any last invalid columns
            for val in xls_df.columns:
                if val is None:
                    to_drop = [None]
                    print("Invalid column in", fil, ":", to_drop, '\n')
                    xls_df.drop(to_drop, inplace=True, axis=1)
                else:
                    continue
            # Collect data.frames in series
            data_df.append(xls_df)
            # Make an overview of the location of data.frames, i.e. their different files
            df_dict_ = {os.path.basename(fil) + '_' + str(a_sheet)[12:-2]: n}
            dfs_dict.update(df_dict_)
            n += 1
    # Load Access Database files as data.frame using the pyodbc driver
    # Takes some time to establish the connection, send sql lines, and get data into data.frames
    # ... better to use a multi-threaded db connection, than serializing all database access
    elif fil[-6:] == ".accdb" and fil in file_search:
        access_path = file_list.get(str(fil))
        # Connect to the database
        conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'
                              r'DBQ=' + access_path + ';')
        # Get blank variable lists to assign them the existing table names in the Access databases
        k_table = select(fil[:-6])
        current_table = k_table[0][1]
        # Fetch all tables in the file
        dbase_tables = conn.cursor().tables()
        for table in dbase_tables:
            if table[3] == 'TABLE':
                current_table.append(table[2])
        # Load each existing table and assign the data to the data.frame
        for item in current_table:
            query1 = "SELECT column_name, data_type FROM information_schema.columns WHERE table_name = {0}".format(item)
            query2 = "SELECT * FROM {0}".format(item)
            # Create a data.frame from each data point, which corresponds to a table in the file
            adb_df = pd.read_sql(query2, conn)
            # Collect data.frames in series
            data_df.append(adb_df)
            # Make an overview of the location of data.frames, i.e. their different files
            df_dict_ = {os.path.basename(fil + "_" + str(item)): n}
            dfs_dict.update(df_dict_)
            n += 1
        # End connection
        conn.close()
    else:
        continue
